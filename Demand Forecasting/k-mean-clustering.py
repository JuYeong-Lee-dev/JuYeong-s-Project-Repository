# =============================================================================
# SECTION 1.  IMPORTS & DATABASE CONNECTION
# =============================================================================

import warnings
warnings.filterwarnings('ignore')

import numpy as np
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import matplotlib.gridspec as gridspec
import plotly.graph_objects as go
from matplotlib.colors import LinearSegmentedColormap
from matplotlib import colors as mcolors
from scipy.stats import linregress
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
from sklearn.decomposition import PCA
from yellowbrick.cluster import KElbowVisualizer, SilhouetteVisualizer
from sklearn.metrics import silhouette_score, calinski_harabasz_score, davies_bouldin_score
from sklearn.cluster import KMeans
from tabulate import tabulate
from collections import Counter

import pandas as pd
import numpy as np 
from sqlalchemy import create_engine, text

from dotenv import load_dotenv
import os
load_dotenv()
engine = create_engine(os.environ['DB_URL'])


# =============================================================================
# SECTION 2.  DATA LOADING
# =============================================================================

input_file = "input_data.xlsx"
origin_table_name = "cleaned"

df = pd.read_excel(input_file)
df.to_sql(f"{origin_table_name}", engine, schema = 'raw', if_exists="replace", index=False)


# =============================================================================
# SECTION 3.  FEATURE ENGINEERING  (SQL)
# =============================================================================

numeric_variables = text("""
DROP TABLE IF EXISTS raw.frequency;
CREATE TABLE raw.frequency AS
WITH filtered_data AS (
  SELECT *
  FROM raw.cleaned
  WHERE "Contract Amount" > 0.5
),
date_range AS (
  SELECT 
    MIN("QTN Date"::DATE) AS min_date,
    MAX("QTN Date"::DATE) AS max_date,
    EXTRACT(YEAR FROM AGE(MAX("QTN Date"::DATE), MIN("QTN Date"::DATE))) * 12 + 
    EXTRACT(MONTH FROM AGE(MAX("QTN Date"::DATE), MIN("QTN Date"::DATE))) + 1 AS total_months
  FROM filtered_data
),
all_months AS (
  SELECT DISTINCT TO_CHAR("QTN Date"::DATE, 'YYYY-MM') AS order_month
  FROM filtered_data
),
all_customers AS (
  SELECT DISTINCT "U-CODE"
  FROM filtered_data
),
customer_month_grid AS (
  SELECT 
    c."U-CODE",
    m.order_month
  FROM all_customers c
  CROSS JOIN all_months m
),
monthly_deliveries AS (
  SELECT
    "U-CODE",
    TO_CHAR("QTN Date"::DATE, 'YYYY-MM') AS order_month,
    COUNT(*) AS deliveries_in_month,
    SUM("Q'ty") AS quantity_in_month
  FROM filtered_data
  GROUP BY "U-CODE", TO_CHAR("QTN Date"::DATE, 'YYYY-MM')
),
complete_monthly_data AS (
  SELECT
    g."U-CODE",
    g.order_month,
    COALESCE(m.deliveries_in_month, 0) AS deliveries_in_month,
    COALESCE(m.quantity_in_month, 0) AS quantity_in_month
  FROM customer_month_grid g
  LEFT JOIN monthly_deliveries m 
    ON g."U-CODE" = m."U-CODE" 
    AND g.order_month = m.order_month
)
SELECT 
  c."U-CODE",
  ROUND(AVG(c.deliveries_in_month), 3) AS avg_monthly_frequency,
  ROUND(AVG(c.quantity_in_month),3) AS avg_monthly_quantity,
  ROUND(
    CASE 
      WHEN AVG(c.deliveries_in_month) > 0
      THEN STDDEV(c.deliveries_in_month) / AVG(c.deliveries_in_month)
      ELSE NULL 
    END, 3
  ) AS coefficient_of_variation
FROM complete_monthly_data c
CROSS JOIN date_range d
GROUP BY c."U-CODE"
ORDER BY c."U-CODE";
        """)

with engine.begin() as conn:
    conn.execute(numeric_variables)


# =============================================================================
# SECTION 4.  LOG TRANSFORMATION
# =============================================================================

df = pd.read_sql_table('frequency', engine, schema = 'raw')

columns = ["avg_monthly_frequency", "avg_monthly_quantity"]

transform = df[columns]

log_data = np.log(transform)

df[columns] = log_data

df.head()


# =============================================================================
# SECTION 5.  OUTLIER DETECTION  (Isolation Forest)
# =============================================================================

model = IsolationForest(contamination=0.05, random_state=0)

df['Outlier_Scores'] = model.fit_predict(df.iloc[:, 1:].to_numpy())

df['Is_Outlier'] = [1 if x == -1 else 0 for x in df['Outlier_Scores']]

df.head()

outlier_percentage = df['Is_Outlier'].value_counts(normalize=True) * 100

plt.figure(figsize=(12, 4))
outlier_percentage.plot(kind='barh', color='#ff6200')

for index, value in enumerate(outlier_percentage):
    plt.text(value, index, f'{value:.2f}%', fontsize=15)

plt.title('Percentage of Inliers and Outliers')
plt.xticks(ticks=np.arange(0, 115, 5))
plt.xlabel('Percentage (%)')
plt.ylabel('Is Outlier')
plt.gca().invert_yaxis()
plt.show()

outliers_data = df[df['Is_Outlier'] == 1]

df_cleaned = df[df['Is_Outlier'] == 0]

df_cleaned = df_cleaned.drop(columns=['Outlier_Scores', 'Is_Outlier'])

df_cleaned.reset_index(drop=True, inplace=True)


# =============================================================================
# SECTION 6.  CORRELATION MATRIX
# =============================================================================

sns.set_style('whitegrid')

corr = df_cleaned.drop(columns=['U-CODE']).corr()

colors = ['#ff6200', '#ffcaa8', 'white', '#ffcaa8', '#ff6200']
my_cmap = LinearSegmentedColormap.from_list('custom_map', colors, N=256)

mask = np.zeros_like(corr)
mask[np.triu_indices_from(mask, k=1)] = True

plt.figure(figsize=(12, 10))
sns.heatmap(corr, mask=mask, cmap=my_cmap, annot=True, center=0, fmt='.2f', linewidths=2)
plt.title('Correlation Matrix', fontsize=14)
plt.show()


# =============================================================================
# SECTION 7.  STANDARDISATION & PCA
# =============================================================================

scaler = StandardScaler()

columns_to_exclude = ['U-CODE']

columns_to_scale = df_cleaned.columns.difference(columns_to_exclude)

df_scaled = df_cleaned.copy()

df_scaled[columns_to_scale] = scaler.fit_transform(df_scaled[columns_to_scale])

df_scaled.head()

transform = df_scaled["avg_monthly_frequency"]
test = df_scaled["avg_monthly_quantity"]

plt.hist(transform, bins = 20)
plt.show()

df_scaled.set_index('U-CODE', inplace=True)
pca = PCA().fit(df_scaled)

explained_variance_ratio = pca.explained_variance_ratio_
cumulative_explained_variance = np.cumsum(explained_variance_ratio)

sns.set(rc={'axes.facecolor': '#fcf0dc'}, style='darkgrid')

plt.figure(figsize=(20, 10))

barplot = sns.barplot(x=list(range(1, len(cumulative_explained_variance) + 1)),
                      y=explained_variance_ratio,
                      color='#fcc36d',
                      alpha=0.8)

lineplot, = plt.plot(range(0, len(cumulative_explained_variance)), cumulative_explained_variance,
                     marker='o', linestyle='--', color='#ff6200', linewidth=2)

plt.xlabel('Number of Components', fontsize=14)
plt.ylabel('Explained Variance', fontsize=14)
plt.title('Cumulative Variance vs. Number of Components', fontsize=18)

plt.xticks(range(0, len(cumulative_explained_variance)))
plt.legend(handles=[barplot.patches[0], lineplot],
           labels=['Explained Variance of Each Component', 'Cumulative Explained Variance'],
           loc=(0.62, 0.1),
           frameon=True,
           framealpha=1.0,  
           edgecolor='#ff6200')  

x_offset = -0.3
y_offset = 0.01
for i, (ev_ratio, cum_ev_ratio) in enumerate(zip(explained_variance_ratio, cumulative_explained_variance)):
    plt.text(i, ev_ratio, f"{ev_ratio:.2f}", ha="center", va="bottom", fontsize=10)
    if i > 0:
        plt.text(i + x_offset, cum_ev_ratio + y_offset, f"{cum_ev_ratio:.2f}", ha="center", va="bottom", fontsize=10)

plt.grid(axis='both')   
plt.show()

pca = PCA(n_components=2)

df_pca = pca.fit_transform(df_scaled)

df_pca = pd.DataFrame(df_pca, columns=['PC'+str(i+1) for i in range(pca.n_components_)])

df_pca.index = df_scaled.index


# =============================================================================
# SECTION 8.  OPTIMAL K SELECTION  (Silhouette Analysis)
# =============================================================================

def silhouette_analysis(df, start_k, stop_k, figsize=(15, 16)):
    """
    Perform Silhouette analysis for a range of k values and visualize the results.
    """
    
    n_k_values = stop_k - start_k + 1
    
    n_rows = (n_k_values + 1) // 2 + 1
    
    plt.figure(figsize=figsize)
    
    grid = gridspec.GridSpec(n_rows, 2, hspace=0.4, wspace=0.3)
    
    first_plot = plt.subplot(grid[0, :])
    
    sns.set_palette(['darkorange'])
    
    silhouette_scores = []
    
    for k in range(start_k, stop_k + 1):
        km = KMeans(n_clusters=k, init='k-means++', n_init=10, max_iter=100, random_state=0)
        km.fit(df)
        labels = km.predict(df)
        score = silhouette_score(df, labels)
        silhouette_scores.append(score)
    
    best_k = start_k + silhouette_scores.index(max(silhouette_scores))
    
    plt.plot(range(start_k, stop_k + 1), silhouette_scores, marker='o')
    plt.xticks(range(start_k, stop_k + 1))
    plt.xlabel('Number of clusters (k)')
    plt.ylabel('Silhouette score')
    plt.title('Average Silhouette Score for Different k Values', fontsize=15)
    
    optimal_k_text = f'The k value with the highest Silhouette score is: {best_k}'
    plt.text(0.7, 0.05, optimal_k_text, fontsize=12, transform=first_plot.transAxes,
             verticalalignment='bottom', horizontalalignment='left', 
             bbox=dict(facecolor='#fcc36d', edgecolor='#ff6200', boxstyle='round, pad=0.5'))
    
    colors = sns.color_palette("bright")
    
    for i in range(start_k, stop_k + 1):    
        km = KMeans(n_clusters=i, init='k-means++', n_init=10, max_iter=100, random_state=0)
        
        subplot_idx = i - start_k
        row_idx = (subplot_idx // 2) + 1
        col_idx = subplot_idx % 2
        
        ax = plt.subplot(grid[row_idx, col_idx])
        
        visualizer = SilhouetteVisualizer(km, colors=colors, ax=ax)
        visualizer.fit(df)
        
        score = silhouette_score(df, km.labels_)
        ax.text(0.97, 0.02, f'Silhouette Score: {score:.3f}', fontsize=12,
                ha='right', transform=ax.transAxes, color='red')
        
        ax.set_title(f'Silhouette Plot for {i} Clusters', fontsize=15)
    
    plt.tight_layout()
    plt.show()


# =============================================================================
# SECTION 9.  K-MEANS CLUSTERING  (k=4)
# =============================================================================

silhouette_analysis(df_pca, 3, 7, figsize=(20, 50))

kmeans = KMeans(n_clusters=4, init='k-means++', n_init=10, max_iter=100, random_state=0)
kmeans.fit(df_pca)

cluster_frequencies = Counter(kmeans.labels_)

label_mapping = {label: new_label for new_label, (label, _) in 
                 enumerate(cluster_frequencies.most_common())}

label_mapping = {label: new_label for new_label, (label, _) in 
                 enumerate(cluster_frequencies.most_common())}

new_labels = np.array([label_mapping[label] for label in kmeans.labels_])

df_cleaned['cluster'] = new_labels

df_pca['cluster'] = new_labels

colors = ['#e8000b', '#1ac938', '#023eff', '#ffc61e']

cluster_0 = df_pca[df_pca['cluster'] == 0]
cluster_1 = df_pca[df_pca['cluster'] == 1]
cluster_2 = df_pca[df_pca['cluster'] == 2]
cluster_3 = df_pca[df_pca['cluster'] == 3]

fig = go.Figure()

fig.add_trace(go.Scatter(x=cluster_0['PC1'], y=cluster_0['PC2'], 
                         mode='markers', marker=dict(color=colors[0], size=8, opacity=0.6), name='Cluster 0'))
fig.add_trace(go.Scatter(x=cluster_1['PC1'], y=cluster_1['PC2'], 
                         mode='markers', marker=dict(color=colors[1], size=8, opacity=0.6), name='Cluster 1'))
fig.add_trace(go.Scatter(x=cluster_2['PC1'], y=cluster_2['PC2'], 
                         mode='markers', marker=dict(color=colors[2], size=8, opacity=0.6), name='Cluster 2'))
fig.add_trace(go.Scatter(x=cluster_3['PC1'], y=cluster_3['PC2'], 
                         mode='markers', marker=dict(color=colors[3], size=8, opacity=0.6), name='Cluster 3'))

fig.update_layout(
    title=dict(text='2D Visualization of Customer Clusters in PCA Space', x=0.5),
    xaxis=dict(gridcolor='white', title='PC1'),
    yaxis=dict(gridcolor='white', title='PC2'),
    width=900,
    height=700,
    plot_bgcolor='#fcf0dc'
)

fig.show()


# =============================================================================
# SECTION 10.  CLUSTER EVALUATION METRICS
# =============================================================================

cluster_percentage = (df_pca['cluster'].value_counts(normalize=True) * 100).reset_index()
cluster_percentage.columns = ['Cluster', 'Percentage']
cluster_percentage.sort_values(by='Cluster', inplace=True)

plt.figure(figsize=(10, 4))
sns.barplot(x='Percentage', y='Cluster', data=cluster_percentage, orient='h', palette=colors)

for index, value in enumerate(cluster_percentage['Percentage']):
    plt.text(value+0.5, index, f'{value:.2f}%')

plt.title('Distribution of Materials Across Clusters', fontsize=14)
plt.xticks(ticks=np.arange(0, 50, 5))
plt.xlabel('Percentage (%)')

plt.show()

num_observations = len(df_pca)

X = df_pca.drop('cluster', axis=1)
clusters = df_pca['cluster']

sil_score = silhouette_score(X, clusters)
calinski_score = calinski_harabasz_score(X, clusters)
davies_score = davies_bouldin_score(X, clusters)

table_data = [
    ["Number of Observations", num_observations],
    ["Silhouette Score", sil_score],
    ["Calinski Harabasz Score", calinski_score],
    ["Davies Bouldin Score", davies_score]
]

print(tabulate(table_data, headers=["Metric", "Value"], tablefmt='pretty'))


# =============================================================================
# SECTION 11.  CLUSTER PROFILING  (Radar Chart)
# =============================================================================

df_customer = df_cleaned.set_index('U-CODE')

scaler = StandardScaler()
columns_to_standardize = df_customer.drop(columns=['cluster']).columns
df_customer_standardized = scaler.fit_transform(df_customer.drop(columns=['cluster']))

df_customer_standardized = pd.DataFrame(df_customer_standardized, 
                                        columns=columns_to_standardize,
                                        index=df_customer.index)
df_customer_standardized['cluster'] = df_customer['cluster'].values

cluster_centroids = df_customer_standardized.groupby('cluster').mean()

def create_radar_chart(ax, angles, data, color, cluster):
    ax.fill(angles, data, color=color, alpha=0.4)
    ax.plot(angles, data, color=color, linewidth=2, linestyle='solid')
    
    ax.set_title(f'Cluster {cluster}', size=20, color=color, y=1.1)

labels = np.array(cluster_centroids.columns)
num_vars = len(labels)

angles = np.linspace(0, 2 * np.pi, num_vars, endpoint=False).tolist()

labels = np.concatenate((labels, [labels[0]]))
angles += angles[:1]

fig, ax = plt.subplots(figsize=(24, 6), subplot_kw=dict(polar=True), nrows=1, ncols=4)

for i, color in enumerate(colors):
    data = cluster_centroids.loc[i].tolist()
    data += data[:1]
    create_radar_chart(ax[i], angles, data, color, i)
    
    ax[i].set_xticks(angles[:-1])
    ax[i].set_xticklabels(labels[:-1], size=10)
    ax[i].grid(color='grey', linewidth=0.5)

plt.tight_layout()
plt.show()


# =============================================================================
# SECTION 12.  GRADE ASSIGNMENT
# =============================================================================

df_grade = df_cleaned.copy()

columns = ["avg_monthly_frequency", "avg_monthly_quantity"]
df_grade[columns] = np.exp(df_grade[columns])

df_grade.head()

grade_mapping = {3: 'D', 2: 'A', 1: 'C', 0: 'B'}

df_grade['Grade'] = df_grade['cluster'].map(grade_mapping)

df_grade = df_grade.drop('cluster', axis = 1)

df_grade.head()


# =============================================================================
# SECTION 13.  COMBINE GRADES & BENCHMARK COMPARISON
# =============================================================================

outliers_data = outliers_data.drop(['Outlier_Scores', 'Is_Outlier'], axis = 1)
outliers_data['Grade'] = 'Outlier'
outliers_data[columns] = np.exp(outliers_data[columns])
df_combined = pd.concat([df_grade, outliers_data], ignore_index=True)

df_combined.head()

import openpyxl

wb = openpyxl.load_workbook('benchmark_raw.xlsx')
ws = wb.active  # or wb['SheetName'] for a specific sheet

max_col = ws.max_column

for col in range(max_col, 6, -1):
    ws.delete_cols(col)

ws.delete_cols(5)
ws.delete_cols(4)
ws.delete_cols(3)
ws.delete_cols(1)

ws['A1'] = 'Grade_Benchmark'
ws['B1'] = 'U-CODE_Benchmark'

wb.save('benchmark_clean.xlsx')

df_combined.to_sql("frequency_combined", engine, schema = 'raw', if_exists = "replace", index = False)
gspm = pd.read_excel('benchmark_clean.xlsx')
gspm.to_sql("gspm", engine, schema = 'raw', if_exists = 'replace', index = False )

compare = text("""
DROP TABLE IF EXISTS raw.compare;
CREATE TABLE raw.compare AS
SELECT a.*, b."Grade_Benchmark"
FROM raw.frequency_combined AS a
LEFT JOIN raw.gspm AS b
ON a."U-CODE" = b."U-CODE_Benchmark"            
               """)

with engine.begin() as conn:
    conn.execute(compare)

df_final = pd.read_sql_table('compare', engine, schema = 'raw')
df_final.loc[df_final['Grade'] == 'Outlier', 'Grade'] = df_final.loc[df_final['Grade'] == 'Outlier', 'Grade_Benchmark']
df_final.to_excel("clustering_results.xlsx", index = False)
